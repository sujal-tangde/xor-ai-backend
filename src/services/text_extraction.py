"""Extract and chunk text from uploaded documents (pdf/docx/xlsx/txt).

Raw text is pulled per file type, then split with LangChain's
``RecursiveCharacterTextSplitter`` into overlapping chunks suitable for
embedding and map/reduce analysis.
"""

from __future__ import annotations

import io
import logging

from langchain_text_splitters import RecursiveCharacterTextSplitter

logger = logging.getLogger(__name__)

CHUNK_SIZE = 1200
CHUNK_OVERLAP = 150

_splitter = RecursiveCharacterTextSplitter(
    chunk_size=CHUNK_SIZE,
    chunk_overlap=CHUNK_OVERLAP,
    separators=["\n\n", "\n", ". ", " ", ""],
)


def _extract_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            logger.exception("Failed to extract a PDF page")
    return "\n\n".join(parts)


def _extract_docx(data: bytes) -> str:
    import docx2txt

    # docx2txt only reads from a path/file-like; a BytesIO works.
    return docx2txt.process(io.BytesIO(data)) or ""


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)


def _extract_txt(data: bytes) -> str:
    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def extract_text(data: bytes, ext: str) -> str:
    """Return the raw text content of a document by extension."""
    ext = ext.lower().lstrip(".")
    if ext == "pdf":
        return _extract_pdf(data)
    if ext == "docx":
        return _extract_docx(data)
    if ext == "xlsx":
        return _extract_xlsx(data)
    if ext == "txt":
        return _extract_txt(data)
    raise ValueError(f"Unsupported document type for text extraction: .{ext}")


def split_text(text: str) -> list[str]:
    """Split arbitrary text into overlapping chunks (empty list if blank)."""
    text = (text or "").strip()
    if not text:
        return []
    return [c for c in _splitter.split_text(text) if c.strip()]


def chunk_document(data: bytes, ext: str) -> list[str]:
    """Extract and split a document into text chunks (empty list if no text)."""
    return split_text(extract_text(data, ext))
